import csv
import io
import json
import os
import zipfile
from datetime import datetime

from flask import Blueprint, request, jsonify, Response
from auth_routes import requiere_auth
from .helpers import _nido, get_db

bp = Blueprint('datos', __name__)


# ── Backup ────────────────────────────────────────────────────────────────────

@bp.route('/api/backup', methods=['GET'])
def api_backup_descargar():
    """Descarga el backup actual como JSON."""
    from flask import current_app
    guardar_backup = current_app.config.get('guardar_backup')
    if guardar_backup:
        guardar_backup()
    backup_path = current_app.config.get('BACKUP_PATH', '')
    if not backup_path or not os.path.exists(backup_path):
        return jsonify({'error': 'no hay backup'}), 404
    with open(backup_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return jsonify(data)


@bp.route('/api/backup', methods=['POST'])
def api_backup_forzar():
    from flask import current_app
    guardar_backup = current_app.config.get('guardar_backup')
    if guardar_backup:
        guardar_backup()
    backup_path = current_app.config.get('BACKUP_PATH', '')
    size = os.path.getsize(backup_path) if backup_path and os.path.exists(backup_path) else 0
    return jsonify({'ok': True, 'size_bytes': size, 'path': backup_path})


# ── GDPR — Exportar y eliminar datos ─────────────────────────────────────────

@bp.route('/api/mis-datos')
def api_mis_datos():
    """Exporta todos los datos del nido actual en JSON (GDPR)."""
    nid = _nido()
    conn = get_db()
    data = {
        'exportado_en': datetime.now().isoformat(),
        'gastos': [dict(r) for r in conn.execute(
            "SELECT * FROM gastos WHERE nido_id=? ORDER BY fecha", (nid,)).fetchall()],
        'ingresos': [dict(r) for r in conn.execute(
            "SELECT * FROM ingresos WHERE nido_id=? ORDER BY fecha", (nid,)).fetchall()],
        'metas': [dict(r) for r in conn.execute(
            "SELECT * FROM metas_ahorro WHERE nido_id=? ORDER BY created_at", (nid,)).fetchall()],
        'presupuestos': [dict(r) for r in conn.execute(
            "SELECT * FROM presupuestos WHERE nido_id=?", (nid,)).fetchall()],
        'categorias': [dict(r) for r in conn.execute(
            "SELECT * FROM categorias_gasto WHERE nido_id=?", (nid,)).fetchall()],
    }
    conn.close()
    return jsonify(data)


@bp.route('/api/eliminar-cuenta', methods=['POST'])
def api_eliminar_cuenta():
    """Elimina permanentemente todos los datos del usuario autenticado (GDPR)."""
    from auth_routes import get_nido_id_from_request
    d = request.get_json() or {}
    confirmacion = d.get('confirmar', '')
    if confirmacion != 'ELIMINAR':
        return jsonify({'error': 'Confirmación incorrecta'}), 400

    firebase_uid = d.get('firebase_uid', '')
    if not firebase_uid:
        return jsonify({'error': 'firebase_uid requerido'}), 400

    conn = get_db()
    fuser = conn.execute(
        "SELECT * FROM firebase_users WHERE firebase_uid=?", (firebase_uid,)
    ).fetchone()
    if not fuser:
        conn.close()
        return jsonify({'error': 'Usuario no encontrado'}), 404

    nid = fuser['nido_id']
    uid_usr = fuser['usuario_id']

    # Comprobar si es el único miembro del nido
    otros = conn.execute(
        "SELECT COUNT(*) as c FROM firebase_users WHERE nido_id=? AND firebase_uid!=?",
        (nid, firebase_uid)
    ).fetchone()['c']

    if otros == 0:
        # Único miembro → eliminar el nido completo
        conn.execute("DELETE FROM aportaciones_meta WHERE meta_id IN (SELECT id FROM metas_ahorro WHERE nido_id=?)", (nid,))
        conn.execute("DELETE FROM gastos WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM ingresos WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM metas_ahorro WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM presupuestos WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM categorias_gasto WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM notificaciones WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM invitaciones WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM firebase_users WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM usuarios WHERE nido_id=?", (nid,))
        conn.execute("DELETE FROM nidos WHERE id=?", (nid,))
    else:
        # Hay otro miembro → solo eliminar este usuario
        conn.execute("DELETE FROM firebase_users WHERE firebase_uid=?", (firebase_uid,))
        conn.execute("UPDATE gastos SET usuario_id=NULL WHERE usuario_id=?", (uid_usr,))
        conn.execute("UPDATE ingresos SET usuario_id=NULL WHERE usuario_id=?", (uid_usr,))
        conn.execute("DELETE FROM notificaciones WHERE usuario_id=?", (uid_usr,))
        conn.execute("DELETE FROM usuarios WHERE id=?", (uid_usr,))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'mensaje': 'Cuenta eliminada correctamente.'})


# ── Exportación de datos (CSV / Excel / PDF) ──────────────────────────────────

def _datos_nido(nid, conn):
    gastos = conn.execute("""
        SELECT g.fecha, g.descripcion, g.importe,
               COALESCE(c.nombre, 'Sin categoría') AS categoria,
               CASE WHEN g.es_fijo THEN 'Sí' ELSE 'No' END AS fijo
        FROM gastos g
        LEFT JOIN categorias_gasto c ON g.categoria_id = c.id
        WHERE g.nido_id=? ORDER BY g.fecha
    """, (nid,)).fetchall()
    ingresos = conn.execute("""
        SELECT fecha, descripcion, importe,
               CASE WHEN es_nomina THEN 'Sí' ELSE 'No' END AS nomina
        FROM ingresos WHERE nido_id=? ORDER BY fecha
    """, (nid,)).fetchall()
    metas = conn.execute("""
        SELECT nombre, importe_objetivo, importe_actual,
               COALESCE(fecha_limite, '-') AS fecha_limite,
               CASE WHEN completada THEN 'Sí' ELSE 'No' END AS completada
        FROM metas_ahorro WHERE nido_id=? ORDER BY created_at
    """, (nid,)).fetchall()
    return {'gastos': gastos, 'ingresos': ingresos, 'metas': metas}


@bp.route('/api/mis-datos/csv')
def api_mis_datos_csv():
    nid = _nido()
    conn = get_db()
    d = _datos_nido(nid, conn)
    conn.close()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nombre, filas in d.items():
            sf = io.StringIO()
            if filas:
                w = csv.DictWriter(sf, fieldnames=filas[0].keys())
                w.writeheader()
                w.writerows(filas)
            zf.writestr(f'{nombre}.csv', sf.getvalue())
    buf.seek(0)
    return Response(buf.getvalue(), mimetype='application/zip',
                    headers={'Content-Disposition': 'attachment; filename=nido_datos.zip'})


@bp.route('/api/mis-datos/excel')
def api_mis_datos_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    nid = _nido()
    conn = get_db()
    d = _datos_nido(nid, conn)
    conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    verde = PatternFill('solid', fgColor='3F5E54')
    font_h = Font(color='F5EDE0', bold=True)

    nombres = {'gastos': 'Gastos', 'ingresos': 'Ingresos', 'metas': 'Metas de ahorro'}
    for key, filas in d.items():
        ws = wb.create_sheet(nombres[key])
        if not filas:
            continue
        cabeceras = list(filas[0].keys())
        for col, cab in enumerate(cabeceras, 1):
            c = ws.cell(row=1, column=col, value=cab.capitalize())
            c.font = font_h
            c.fill = verde
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
        for r, fila in enumerate(filas, 2):
            for col, val in enumerate(fila.values(), 1):
                ws.cell(row=r, column=col, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=nido_datos.xlsx'})


@bp.route('/api/mis-datos/pdf')
def api_mis_datos_pdf():
    from fpdf import FPDF
    nid = _nido()
    conn = get_db()
    d = _datos_nido(nid, conn)
    conn.close()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Cabecera
    pdf.set_fill_color(63, 94, 84)
    pdf.set_text_color(245, 237, 224)
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 14, 'Nido  -  Resumen de datos', ln=True, align='C', fill=True)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 7, f'Exportado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}',
             ln=True, align='C', fill=True)
    pdf.ln(6)

    secciones = [
        ('Gastos',         d['gastos'],   ['fecha', 'descripcion', 'importe', 'categoria', 'fijo']),
        ('Ingresos',       d['ingresos'], ['fecha', 'descripcion', 'importe', 'nomina']),
        ('Metas de ahorro',d['metas'],    ['nombre', 'importe_objetivo', 'importe_actual', 'fecha_limite', 'completada']),
    ]
    col_w = {5: [28, 72, 22, 38, 15], 4: [32, 90, 22, 18], 2: [50, 100, 35], 3: [35, 100, 25, 30]}

    for titulo, filas, cols in secciones:
        pdf.set_text_color(63, 94, 84)
        pdf.set_font('Helvetica', 'B', 13)
        pdf.cell(0, 9, titulo, ln=True)
        pdf.set_draw_color(63, 94, 84)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(2)

        if not filas:
            pdf.set_text_color(150, 150, 150)
            pdf.set_font('Helvetica', 'I', 9)
            pdf.cell(0, 7, 'Sin datos', ln=True)
        else:
            anchos = col_w.get(len(cols), [int(190 / len(cols))] * len(cols))
            pdf.set_fill_color(220, 230, 225)
            pdf.set_text_color(40, 40, 40)
            pdf.set_font('Helvetica', 'B', 9)
            for i, col in enumerate(cols):
                pdf.cell(anchos[i], 7, col.capitalize(), border=1, fill=True)
            pdf.ln()
            pdf.set_font('Helvetica', '', 8)
            for fila in filas:
                for i, col in enumerate(cols):
                    val = str(fila[col] or '')
                    pdf.cell(anchos[i], 6, val[:40], border=1)
                pdf.ln()
        pdf.ln(5)

    pdf.set_text_color(150, 150, 150)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 6, 'Generado por Nido by Kreatia  ·  kreatia3d@gmail.com', align='C')

    buf = io.BytesIO(pdf.output())
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename=nido_datos.pdf'})
